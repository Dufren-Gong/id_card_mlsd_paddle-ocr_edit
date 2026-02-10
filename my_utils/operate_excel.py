import pandas as pd
import os, json, shutil, re
from natsort import natsorted
import numpy as np
from my_utils.Traditional_to_Simplified_Chinese import fan_to_jian
from openpyxl import load_workbook
from openpyxl.styles import Font
from my_utils import check
from my_utils.utils import find_in_catch_pic, find_png_by_name_fast

sheet_names = ['开单', '拿货', '转单', '年费', '补单', '补卡', '退单']
catch_columns = ['性别', '归属地', '民族', '出生日期', '身份证号码', '住址']

column_names = dict(姓名=check.check_name,
                    联系电话=check.check_phone_number,
                    提取金额=check.check_cash,
                    独立经销商卡号=check.check_sail_card_id,
                    货单号码=check.check_sail_id,
                    开单日期=check.check_open_date,
                    剩余货值=check.check_cash,
                    地址=check.check_address)

column_names_after = dict(性别=check.check_sex,
                    民族=check.check_nation,
                    出生日期=check.check_birth,
                    身份证号码=check.check_id,
                    住址=check.check_address)

class People_Info():
    def __init__(self,
                 raw_index_in_excel,
                 name:str = '',
                 telephone:str = '',
                 sex:str = '',
                 birth:str = '',
                 id:str = '',
                 pay:int = 0,
                 buy:int = 0,
                 address:str = '',
                 native:str = '',
                 regin:str = '',
                 extract = '10000',
                 annual_fee:int = 0,
                 sail_card_id = '0',
                 sail_id = '',
                 open_date = '',
                 left_money = '0') -> None:
        self.raw_index_in_excel = raw_index_in_excel
        self.name = name
        self.telephone = telephone
        self.sex = sex
        self.birth = birth
        self.id = id
        self.pay = str(pay)
        self.buy = str(buy)
        self.extract = extract
        self.address = address
        self.native = native
        self.regin = regin
        self.annual_fee = str(annual_fee)
        self.sail_card_id = sail_card_id
        self.sail_id = sail_id
        self.open_date = open_date
        self.left_money = left_money

class Pair():
    def __init__(self,
                 client:People_Info,
                 entrusted:People_Info,
                 beiweituo:People_Info = None) -> None:
        self.client = client
        self.entrusted = entrusted
        self.beiweituo = beiweituo

    def swap_client_and_beiweituo(self):
        self.client, self.beiweituo = self.beiweituo, self.client

    def swap_client_and_entrusted(self):
        self.client, self.entrusted = self.entrusted, self.client

    def swap_entrusted_and_beiweituo(self):
        self.entrusted, self.beiweituo = self.beiweituo, self.entrusted

def load_json_data(path):
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as reader:
            line = reader.readline()
        line = line.strip()
        info = json.loads(line)
    #如果找不到这个单子的信息
    else:
        info_name = os.path.basename(path).rsplit('.', maxsplit = 1)[0]
        cache_path = os.path.join('模版', '高频照片', info_name)
        #如果在cache里面找到了
        if os.path.exists(cache_path + '.info') and os.path.exists(cache_path + '.png') and os.path.exists(cache_path + '反.png'):
            with open(cache_path + '.info', 'r', encoding='utf-8') as reader:
                line = reader.readline()
            line = line.strip()
            info = json.loads(line)
            shutil.copy(cache_path + '.png', os.path.join(os.path.dirname(path), info_name + '.png'))
            shutil.copy(cache_path + '.info', os.path.join(os.path.dirname(path), info_name + '.info'))
            shutil.copy(cache_path + '反.png', os.path.join(os.path.dirname(path), info_name + '反.png'))
        else:
            cache_path = find_in_catch_pic(info_name, './模版/缓存照片', 2)
            if cache_path != '':
                with open(cache_path + '.info', 'r', encoding='utf-8') as reader:
                    line = reader.readline()
                line = line.strip()
                info = json.loads(line)
                shutil.copy(cache_path + '.png', os.path.join(os.path.dirname(path), info_name + '.png'))
                shutil.copy(cache_path + '.info', os.path.join(os.path.dirname(path), info_name + '.info'))
                shutil.copy(cache_path + '反.png', os.path.join(os.path.dirname(path), info_name + '反.png'))
            else:
                info = None
    return info

def get_info(index, excel_row, info_json, add_flag=False, add_card_id_flag=False):
    try:
        assert excel_row['姓名'] != '' and not pd.isna(excel_row['姓名']) and isinstance(excel_row['姓名'], str)
        assert excel_row['联系电话'] != '' and not pd.isna(excel_row['联系电话'])
        if info_json['籍贯'] == '香港':
            assert excel_row['地址'] != '' and not pd.isna(excel_row['地址'])
            info_json['住址'] = excel_row['地址']
            info_json['民族'] = '无'
        else:
            if isinstance(excel_row['地址'], str) and excel_row['地址'].replace(' ', '') != '':
                info_json['住址'] = excel_row['地址']
        if add_flag:
            assert excel_row['提取金额'] != '' and not pd.isna(excel_row['提取金额'])
        try:
            id = excel_row['独立经销商卡号']
        except:
            id = ''
        if add_card_id_flag:
            assert id != '' and not pd.isna(id)
        else:
            id = ''
    except:
        return None
    try:
        return People_Info(index, excel_row['姓名'],
                        excel_row['联系电话'],
                        info_json['性别'],
                        info_json['出生日期'],
                        info_json['身份证号码'],
                        check.moren_pay,
                        check.moren_buy,
                        info_json['住址'],
                        info_json['籍贯'],
                        info_json['民族'],
                        excel_row['提取金额'],
                        check.annual_fee,
                        id)
    except:
        return None

def get_nahuo_info(index, excel_row, info_json, add_flag = False):
    try:
        assert excel_row['姓名'] != '' and not pd.isna(excel_row['姓名']) and isinstance(excel_row['姓名'], str)
        assert excel_row['联系电话'] != '' and not pd.isna(excel_row['联系电话'])
        assert excel_row['独立经销商卡号'] != '' and not pd.isna(excel_row['独立经销商卡号'])
        if info_json['籍贯'] == '香港':
            assert excel_row['地址'] != '' and not pd.isna(excel_row['地址'])
            info_json['住址'] = excel_row['地址']
            info_json['民族'] = '无'
        else:
            if isinstance(excel_row['地址'], str) and excel_row['地址'].replace(' ', '') != '':
                info_json['住址'] = excel_row['地址']
        if add_flag:
            assert excel_row['货单号码'] != '' and not pd.isna(excel_row['货单号码'])
            assert excel_row['提取金额'] != '' and not pd.isna(excel_row['提取金额'])
    except:
        return None
    try:
        return People_Info(index, excel_row['姓名'],
                        excel_row['联系电话'],
                        info_json['性别'],
                        info_json['出生日期'],
                        info_json['身份证号码'],
                        '0',
                        '0',
                        info_json['住址'],
                        info_json['籍贯'],
                        info_json['民族'],
                        excel_row['提取金额'],
                        '0',
                        excel_row['独立经销商卡号'],
                        excel_row['货单号码'])
    except:
        return None

def get_zhuandan_info(index, excel_row, info_json, add_flag=False, end_flag=False):
    try:
        assert excel_row['姓名'] != '' and not pd.isna(excel_row['姓名']) and isinstance(excel_row['姓名'], str)
        assert excel_row['联系电话'] != '' and not pd.isna(excel_row['联系电话'])
        if info_json['籍贯'] == '香港':
            assert excel_row['地址'] != '' and not pd.isna(excel_row['地址'])
            info_json['住址'] = excel_row['地址']
            info_json['民族'] = '无'
        else:
            if isinstance(excel_row['地址'], str) and excel_row['地址'].replace(' ', '') != '':
                info_json['住址'] = excel_row['地址']
        if add_flag:
            assert excel_row['货单号码'] != '' and not pd.isna(excel_row['货单号码'])
            assert excel_row['剩余货值'] != '' and not pd.isna(excel_row['剩余货值'])
            assert excel_row['开单日期'] != '' and not pd.isna(excel_row['开单日期'])
            # excel_row['开单日期'] = manage_data(excel_row['开单日期'])
        if not end_flag:
            assert excel_row['独立经销商卡号'] != '' and not pd.isna(excel_row['独立经销商卡号'])
    except:
        return None
    try:
        return People_Info(index, excel_row['姓名'],
                       excel_row['联系电话'],
                       info_json['性别'],
                       info_json['出生日期'],
                       info_json['身份证号码'],
                       '0',
                       '0',
                       info_json['住址'],
                       info_json['籍贯'],
                       info_json['民族'],
                       '0',
                       check.annual_fee,
                       excel_row['独立经销商卡号'],
                       excel_row['货单号码'],
                       excel_row['开单日期'],
                       excel_row['剩余货值'])
    except:
        return None
    
def get_budan_info(index, excel_row, info_json, add_flag=False):
    try:
        assert excel_row['姓名'] != '' and not pd.isna(excel_row['姓名']) and isinstance(excel_row['姓名'], str)
        assert excel_row['联系电话'] != '' and not pd.isna(excel_row['联系电话'])
        assert excel_row['独立经销商卡号'] != '' and not pd.isna(excel_row['独立经销商卡号'])
        if info_json['籍贯'] == '香港':
            assert excel_row['地址'] != '' and not pd.isna(excel_row['地址'])
            info_json['住址'] = excel_row['地址']
            info_json['民族'] = '无'
        else:
            if isinstance(excel_row['地址'], str) and excel_row['地址'].replace(' ', '') != '':
                info_json['住址'] = excel_row['地址']
        if add_flag:
            assert excel_row['货单号码'] != '' and not pd.isna(excel_row['货单号码'])
    except:
        return None
    try:
        return People_Info(index, excel_row['姓名'],
                       excel_row['联系电话'],
                       info_json['性别'],
                       info_json['出生日期'],
                       info_json['身份证号码'],
                       '0',
                       '0',
                       info_json['住址'],
                       info_json['籍贯'],
                       info_json['民族'],
                       '0',
                       check.annual_fee,
                       excel_row['独立经销商卡号'],
                       excel_row['货单号码'])
    except:
        return None

def get_buka_info(index, excel_row, info_json, add_flag = False):
    try:
        assert excel_row['姓名'] != '' and not pd.isna(excel_row['姓名']) and isinstance(excel_row['姓名'], str)
        assert excel_row['联系电话'] != '' and not pd.isna(excel_row['联系电话'])
        assert excel_row['独立经销商卡号'] != '' and not pd.isna(excel_row['独立经销商卡号'])
        if info_json['籍贯'] == '香港':
            assert excel_row['地址'] != '' and not pd.isna(excel_row['地址'])
            info_json['住址'] = excel_row['地址']
            info_json['民族'] = '无'
        else:
            if isinstance(excel_row['地址'], str) and excel_row['地址'].replace(' ', '') != '':
                info_json['住址'] = excel_row['地址']
        if add_flag:
            assert excel_row['货单号码'] != '' and not pd.isna(excel_row['货单号码'])
    except:
        return None
    try:
        return People_Info(index, excel_row['姓名'],
                       excel_row['联系电话'],
                       info_json['性别'],
                       info_json['出生日期'],
                       info_json['身份证号码'],
                       '0',
                       '0',
                       info_json['住址'],
                       info_json['籍贯'],
                       info_json['民族'],
                       '0',
                       check.annual_fee,
                       excel_row['独立经销商卡号'],
                       excel_row['货单号码'])
    except:
        return None

def get_tuidan_info(index, excel_row, info_json, add_flag = False):
    try:
        assert excel_row['姓名'] != '' and not pd.isna(excel_row['姓名']) and isinstance(excel_row['姓名'], str)
        assert excel_row['联系电话'] != '' and not pd.isna(excel_row['联系电话'])
        if info_json['籍贯'] == '香港':
            assert excel_row['地址'] != '' and not pd.isna(excel_row['地址'])
            info_json['住址'] = excel_row['地址']
            info_json['民族'] = '无'
        else:
            if isinstance(excel_row['地址'], str) and excel_row['地址'].replace(' ', '') != '':
                info_json['住址'] = excel_row['地址']
        if add_flag:
            assert excel_row['独立经销商卡号'] != '' and not pd.isna(excel_row['独立经销商卡号'])
            assert excel_row['货单号码'] != '' and not pd.isna(excel_row['货单号码'])
    except:
        return None
    try:
        return People_Info(index, excel_row['姓名'],
                       excel_row['联系电话'],
                       info_json['性别'],
                       info_json['出生日期'],
                       info_json['身份证号码'],
                       '0',
                       '0',
                       info_json['住址'],
                       info_json['籍贯'],
                       info_json['民族'],
                       '0',
                       check.annual_fee,
                       excel_row['独立经销商卡号'],
                       excel_row['货单号码'])
    except:
        return None

def get_nianfei_info(index, excel_row, info_json):
    try:
        assert excel_row['姓名'] != '' and not pd.isna(excel_row['姓名']) and isinstance(excel_row['姓名'], str)
        assert excel_row['联系电话'] != '' and not pd.isna(excel_row['联系电话'])
        assert excel_row['独立经销商卡号'] != '' and not pd.isna(excel_row['独立经销商卡号'])
        if info_json['籍贯'] == '香港':
            assert excel_row['地址'] != '' and not pd.isna(excel_row['地址'])
            info_json['住址'] = excel_row['地址']
            info_json['民族'] = '无'
        else:
            if isinstance(excel_row['地址'], str) and excel_row['地址'].replace(' ', '') != '':
                info_json['住址'] = excel_row['地址']
    except:
        return None
    try:
        return People_Info(index, excel_row['姓名'],
                       excel_row['联系电话'],
                       info_json['性别'],
                       info_json['出生日期'],
                       info_json['身份证号码'],
                       '0',
                       '0',
                       info_json['住址'],
                       info_json['籍贯'],
                       info_json['民族'],
                       '0',
                       check.annual_fee,
                       excel_row['独立经销商卡号'])
    except:
        return None

def get_zhuandan_pairs(df, base_path):
    pairs = []
    errors = []
    record_flag = True
    obj1 = None
    obj2 = None
    obj3 = None
    for index, row in df.iterrows():
        if index % 4 != 3:
            name = row['姓名']
            or_name = fan_to_jian(row['姓名'])
            info_path = os.path.join(base_path, f"{name}.info")
            info = load_json_data(info_path)
            #如果没读取到这个人的信息
            if info == None:
                #简体找不到可能是内地人名字含有繁体
                info_path = os.path.join(base_path, f"{or_name}.info")
                info = load_json_data(info_path)
                if info == None:
                    record_flag = False
                    if not pd.isna(name):
                        errors.append(name)
                    continue
                else:
                    name = or_name
            #转让
            if record_flag:
                if index % 4 == 0:
                    obj1 = get_zhuandan_info(index, row, info, add_flag=True)
                    if obj1 == None:
                        errors.append(name)
                #被转让
                elif index % 4 == 1:
                    obj2 = get_zhuandan_info(index, row, info, end_flag=True)
                    if obj2 == None:
                        errors.append(name)
                #被委托
                elif index % 4 == 2:
                    obj3 = get_zhuandan_info(index, row, info, end_flag=True)
                    if obj3 == None:
                        errors.append(name)
        else:
            record_flag = True
        if obj1 != None and obj2 != None and obj3 != None:
            pair_obj = Pair(obj1, obj2, obj3)
            pairs.append(pair_obj)
            obj1 = None
            obj2 = None
            obj3 = None
    return pairs, errors

def get_kaidan_pairs(df, base_path):
    peoples = 2
    errors = []
    pairs = []
    objs = []
    for index, row in df.iterrows():
        name = row['姓名']
        or_name = fan_to_jian(row['姓名'])
        #如果不是空行
        if isinstance(name, str):
            info_path = os.path.join(base_path, f"{name}.info")
            info = load_json_data(info_path)
            #如果没读取到这个人的信息
            if info == None:
                #简体找不到可能是内地人名字含有繁体
                info_path = os.path.join(base_path, f"{or_name}.info")
                info = load_json_data(info_path)
                if info == None:
                    objs.append(None)
                    if not pd.isna(name):
                        errors.append(name)
                    continue
                else:
                    name = or_name
            add_flag = not pd.isna(df.iloc[index + 1]['姓名'])
            try:
                if index == 0:
                    if not pd.isna(df.iloc[index + 1]['独立经销商卡号']):
                        add_card_id_flag = True
                    else:
                        add_card_id_flag = False
                else:
                    if (pd.isna(df.iloc[index - 1]['姓名']) and not pd.isna(df.iloc[index + 1]['独立经销商卡号'])) or \
                    (pd.isna(df.iloc[index + 1]['姓名']) and not pd.isna(df.iloc[index - 1]['独立经销商卡号'])):
                        add_card_id_flag = True
                    else:
                        add_card_id_flag = False
            except:
                add_card_id_flag = False
            obj_temp = get_info(index, row, info, add_flag, add_card_id_flag)
            objs.append(obj_temp)
            if obj_temp == None:
                errors.append(name)
        #如果检测到空行
        elif pd.isna(name):
            #如果行数够了
            if len(objs) >= peoples:
                #如果被委托人信息没错
                if objs[-1] != None:
                    for i in range(len(objs) - 1):
                        #如果该被委托人信息正确
                        if objs[i] != None:
                            pairs.append(Pair(objs[i], objs[-1]))
            objs = []
    return pairs, errors

def get_nahuo_pairs(df, base_path):
    peoples = 2
    errors = []
    pairs = []
    objs = []
    for index, row in df.iterrows():
        name = row['姓名']
        or_name = fan_to_jian(row['姓名'])
        #如果不是空行
        if isinstance(name, str):
            info_path = os.path.join(base_path, f"{name}.info")
            info = load_json_data(info_path)
            #如果没读取到这个人的信息
            if info == None:
                #简体找不到可能是内地人名字含有繁体
                info_path = os.path.join(base_path, f"{or_name}.info")
                info = load_json_data(info_path)
                if info == None:
                    objs.append(None)
                    if not pd.isna(name):
                        errors.append(name)
                    continue
                else:
                    name = or_name
            add_flag = not pd.isna(df.iloc[index + 1]['姓名'])
            obj_temp = get_nahuo_info(index, row, info, add_flag)
            objs.append(obj_temp)
            if obj_temp == None:
                errors.append(name)
        #如果检测到空行
        elif pd.isna(name):
            #如果行数够了
            if len(objs) >= peoples:
                #如果被委托人信息没错
                if objs[-1] != None:
                    for i in range(len(objs) - 1):
                        #如果该被委托人信息正确
                        if objs[i] != None:
                            pairs.append(Pair(objs[i], objs[-1]))
            objs = []
    return pairs, errors

def get_nianfei_pairs(df, base_path):
    peoples = 2
    errors = []
    pairs = []
    objs = []
    for index, row in df.iterrows():
        name = row['姓名']
        or_name = fan_to_jian(row['姓名'])
        #如果不是空行
        if isinstance(name, str):
            info_path = os.path.join(base_path, f"{name}.info")
            info = load_json_data(info_path)
            #如果没读取到这个人的信息
            if info == None:
                #简体找不到可能是内地人名字含有繁体
                info_path = os.path.join(base_path, f"{or_name}.info")
                info = load_json_data(info_path)
                if info == None:
                    objs.append(None)
                    if not pd.isna(name):
                        errors.append(name)
                    continue
                else:
                    name = or_name
            obj_temp = get_nianfei_info(index, row, info)
            if obj_temp == None:
                errors.append(name)
            objs.append(obj_temp)
        #如果检测到空行
        elif pd.isna(name):
            #如果行数够了
            if len(objs) >= peoples:
                #如果被委托人信息没错
                if objs[-1] != None:
                    for i in range(len(objs) - 1):
                        #如果该被委托人信息正确
                        if objs[i] != None:
                            pairs.append(Pair(objs[i], objs[-1]))
            objs = []
    return pairs, errors

def get_budan_pairs(df, base_path):
    peoples = 2
    errors = []
    pairs = []
    objs = []
    for index, row in df.iterrows():
        name = row['姓名']
        or_name = fan_to_jian(row['姓名'])
        #如果不是空行
        if isinstance(name, str):
            info_path = os.path.join(base_path, f"{name}.info")
            info = load_json_data(info_path)
            #如果没读取到这个人的信息
            if info == None:
                #简体找不到可能是内地人名字含有繁体
                info_path = os.path.join(base_path, f"{or_name}.info")
                info = load_json_data(info_path)
                if info == None:
                    objs.append(None)
                    if not pd.isna(name):
                        errors.append(name)
                    continue
                else:
                    name = or_name
            add_flag = not pd.isna(df.iloc[index + 1]['姓名'])
            obj_temp = get_budan_info(index, row, info, add_flag)
            if obj_temp == None:
                errors.append(name)
            objs.append(obj_temp)
        #如果检测到空行
        elif pd.isna(name):
            #如果行数够了
            if len(objs) >= peoples:
                #如果被委托人信息没错
                if objs[-1] != None:
                    for i in range(len(objs) - 1):
                        #如果该被委托人信息正确
                        if objs[i] != None:
                            pairs.append(Pair(objs[i], objs[-1]))
            objs = []
    return pairs, errors

def get_buka_pairs(df, base_path):
    peoples = 2
    errors = []
    pairs = []
    objs = []
    for index, row in df.iterrows():
        name = row['姓名']
        or_name = fan_to_jian(row['姓名'])
        #如果不是空行
        if isinstance(name, str):
            info_path = os.path.join(base_path, f"{name}.info")
            info = load_json_data(info_path)
            #如果没读取到这个人的信息
            if info == None:
                #简体找不到可能是内地人名字含有繁体
                info_path = os.path.join(base_path, f"{or_name}.info")
                info = load_json_data(info_path)
                if info == None:
                    objs.append(None)
                    if not pd.isna(name):
                        errors.append(name)
                    continue
                else:
                    name = or_name
            add_flag = not pd.isna(df.iloc[index + 1]['姓名'])
            obj_temp = get_buka_info(index, row, info, add_flag)
            if obj_temp == None:
                errors.append(name)
            objs.append(obj_temp)
        #如果检测到空行
        elif pd.isna(name):
            #如果行数够了
            if len(objs) >= peoples:
                #如果被委托人信息没错
                if objs[-1] != None:
                    for i in range(len(objs) - 1):
                        #如果该被委托人信息正确
                        if objs[i] != None:
                            pairs.append(Pair(objs[i], objs[-1]))
            objs = []
    return pairs, errors

def get_tuidan_pairs(df, base_path):
    peoples = 2
    errors = []
    pairs = []
    objs = []
    for index, row in df.iterrows():
        name = row['姓名']
        or_name = fan_to_jian(row['姓名'])
        #如果不是空行
        if isinstance(name, str):
            info_path = os.path.join(base_path, f"{name}.info")
            info = load_json_data(info_path)
            #如果没读取到这个人的信息
            if info == None:
                #简体找不到可能是内地人名字含有繁体
                info_path = os.path.join(base_path, f"{or_name}.info")
                info = load_json_data(info_path)
                if info == None:
                    objs.append(None)
                    if not pd.isna(name):
                        errors.append(name)
                    continue
                else:
                    name = or_name
            add_flag = not pd.isna(df.iloc[index + 1]['姓名'])
            obj_temp = get_tuidan_info(index, row, info, add_flag)
            if obj_temp == None:
                errors.append(name)
            objs.append(obj_temp)
        #如果检测到空行
        elif pd.isna(name):
            #如果行数够了
            if len(objs) >= peoples:
                #如果被委托人信息没错
                if objs[-1] != None:
                    for i in range(len(objs) - 1):
                        #如果该被委托人信息正确
                        if objs[i] != None:
                            pairs.append(Pair(objs[i], objs[-1]))
            objs = []
    return pairs, errors

def read_sheets(file_path, sheet_name = None, add_nan_line = True) -> pd.DataFrame:
    #sheet_name = None for all
    #{sheet_name: data}
    df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
    if add_nan_line and sheet_name != None:
        df.loc[len(df)] = np.nan
    return df

def get_column_letter(ws, c_names):
    column_letters = []
    for column in c_names:
        for cell in ws[1]:  # 第一行的单元格
            if cell.value == column:
                column_letters.append(cell.column_letter)
                break
    return column_letters

def fill_information(obj_pairs:Pair, save_path, cue, cache_all_flag = False):
    # 检查 i 是否在 DataFrame 的范围内
    cache_objs = [obj_pairs.client, obj_pairs.entrusted]
    if cache_all_flag:
        cache_objs.append(obj_pairs.beiweituo)
    wb = load_workbook(save_path)
    ws = wb[cue]
    
    column_letters = get_column_letter(ws, catch_columns)

    for obj in cache_objs:
        map_obj = [obj.sex, obj.native, obj.regin, obj.birth, obj.id, obj.address]
        index = obj.raw_index_in_excel
        for index_letter, letter in enumerate(column_letters):
            ws[f"{letter}{index + 2}"] = ''
            ws[f"{letter}{index + 2}"] = map_obj[index_letter]

    wb.save(save_path)

def read_openpyxl_by_str(ws):
    for row in ws.iter_rows():
        for cell in row:
            # 将单元格的值转换为字符串
            cell.value = str(cell.value) if cell.value is not None else None
    return ws

def map_info(ws, template_column_info:str, i, check_cloumns, column_letters):
    infos_t = re.sub(r' +', ' ', template_column_info)
    infos_t = infos_t.replace(':', '：')
    infos_t = infos_t.split('\n')
    infos = []
    for tt in infos_t:
        if tt != '' and tt != ' ':
            infos.append(tt)
    for name_index, j in enumerate(check_cloumns):
        start_index = -1
        for index, info in enumerate(infos):
            if j + '：' in info or j + ' ：' in info:
                start_index = index
                break
        if start_index == -1:
            continue

        temp_info_arr = [info]
        if start_index != len(infos) - 1:
            for k in range(start_index + 1, len(infos)):
                if "：" in infos[k]:
                    break
                else:
                    temp_info_arr.append(infos[k])
        
        result_str = ''
        for m in temp_info_arr:
            infos.remove(m)
            if "：" in m:
                a = m.split('：', maxsplit=1)[-1]
                result_str += a.strip().replace(' ', '')
            else:
                result_str += m.strip().replace(' ', '')
        if result_str != '':
            name_t = column_letters[name_index]
            # if j == '姓名':
            #     result_str = fan_to_jian(result_str)
            ws[f"{name_t}{i}"].value = result_str
    return ws

def chech_and_add(arr_t: list, i):
    if i not in arr_t:
        arr_t.append(i)
    return arr_t

def check_excel(file_path, pic_floader, sheet_name = None, search_extra_floader = (False, '')):
    passed_flag = True
    error_rows = []
    # 设置绿色字体
    red_font = Font(color="FF0000")  # 红色字体
    black_font = Font(color="000000")  # 黑色字体
    blue_font = Font(color="0000FF")  # 蓝色字体
    #高频信息里面找
    cache_path = os.path.join('模版', '高频照片')
    #照片缓存信息里面找
    cache_date_floader = os.path.join('模版', '缓存照片')
    all_floaders_catch = natsorted(os.listdir(cache_date_floader), reverse=True)
    all_floaders_catch = [i for i in all_floaders_catch if os.path.isdir(os.path.join(cache_date_floader, i))]
    if '.DS_Store' in all_floaders_catch:
        all_floaders_catch.remove('.DS_Store')
    sheets = []
    if sheet_name == None:
        sheets = sheet_names
    elif isinstance(sheet_name, str):
        if sheet_name in sheet_names:
            sheets = [sheet_name]
    elif isinstance(sheet_name, list):
        sheets = [i for i in sheet_name if i in sheet_names]
    c_names = column_names.keys()
    wb = load_workbook(file_path)
    extra_searched = []
    if search_extra_floader[0]:
        save_path = './查找到的照片'
        if os.path.exists(save_path):
            shutil.rmtree(save_path)
        os.makedirs(save_path, exist_ok=True)
    for sheet in sheets:
        ws = wb[sheet]
        ws = read_openpyxl_by_str(ws)
        row_count = ws.max_row
        # 遍历所有单元格并设置字体颜色为黑色
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = black_font  # 设置字体颜色为黑色
        check_cloumns = [i for i in [cell.value for cell in ws[1]] if i in c_names]
        column_letters = get_column_letter(ws, check_cloumns + ['模版信息'])
        address_column_index = check_cloumns.index('地址')
        address_column_letters = column_letters[address_column_index]
        template_column_letters = column_letters[-1]
        column_letters = column_letters[:-1]
        #将模版信息展平
        for i in range(2, row_count + 1):
            template_column_info = ws[f"{template_column_letters}{i}"].value
            if not pd.isna(template_column_info):
                ws = map_info(ws, template_column_info, i, check_cloumns, column_letters)

        for index, letter in enumerate(column_letters):
            c_name = check_cloumns[index]
            for i in range(2, row_count + 1):
                value = ws[f"{letter}{i}"].value
                if not pd.isna(value):
                    if value.strip().replace(' ', '') != '':
                        check_r = column_names[c_name](value)
                        if c_name != '姓名':
                            if check_r == None:
                                ws[f"{letter}{i}"].font = red_font
                                passed_flag = False
                                error_rows = chech_and_add(error_rows, i)
                            else:
                                ws[f"{letter}{i}"] = check_r
                        #如果是检查姓名，跟地址一起检查
                        else:
                            name_passed_flag = False
                            found_path = ''
                            found_names = [check_r, fan_to_jian(check_r)]
                            for check_name_index, check_r_name in enumerate(found_names):
                                if name_passed_flag:
                                    break
                                else:
                                    try:
                                        assert os.path.exists(os.path.join(pic_floader, f'{check_r_name}.png'))
                                        assert os.path.exists(os.path.join(pic_floader, f'{check_r_name}反.png'))
                                        assert os.path.exists(os.path.join(pic_floader, f'{check_r_name}.info'))
                                        found_path = pic_floader
                                        check_r = check_r_name
                                        name_passed_flag = True
                                        break
                                    except:
                                        try:
                                            assert os.path.exists(os.path.join(cache_path, f'{check_r_name}.png'))
                                            assert os.path.exists(os.path.join(cache_path, f'{check_r_name}反.png'))
                                            assert os.path.exists(os.path.join(cache_path, f'{check_r_name}.info'))
                                            found_path = cache_path
                                            check_r = check_r_name
                                            name_passed_flag = True
                                            break
                                        except:
                                            if len(all_floaders_catch) != 0:
                                                for check_date_floader in all_floaders_catch:
                                                    try:
                                                        assert os.path.exists(os.path.join(cache_date_floader, check_date_floader, f'{check_r_name}.png'))
                                                        assert os.path.exists(os.path.join(cache_date_floader, check_date_floader, f'{check_r_name}反.png'))
                                                        assert os.path.exists(os.path.join(cache_date_floader, check_date_floader, f'{check_r_name}.info'))
                                                        name_passed_flag = True
                                                        check_r = check_r_name
                                                        found_path = os.path.join(cache_date_floader, check_date_floader)
                                                        break
                                                    except:
                                                        if check_name_index == 1:
                                                            name_passed_flag = False
                                                            passed_flag = False
                                                            error_rows = chech_and_add(error_rows, i)
                                                            ws[f"{letter}{i}"].font = red_font
                                            else:
                                                if check_name_index == 1:
                                                    name_passed_flag = False
                                                    passed_flag = False
                                                    error_rows = chech_and_add(error_rows, i)
                                                    ws[f"{letter}{i}"].font = red_font
                            if not name_passed_flag and search_extra_floader[0]:
                                path = find_png_by_name_fast(search_extra_floader[1], found_names, '.png')
                                if path != None:
                                    this_searched_name = os.path.basename(path).split('.', maxsplit=1)[0]
                                    all_path_root = os.path.dirname(path)
                                    if os.path.exists(os.path.join(all_path_root, f'{this_searched_name}反.png')) and os.path.exists(os.path.join(all_path_root, f'{this_searched_name}.info')):
                                        extra_searched = chech_and_add(extra_searched, this_searched_name)
                                        ws[f"{letter}{i}"].font = blue_font
                                        passed_flag = False
                                        shutil.copy(path, os.path.join(save_path, os.path.basename(path)))
                                        shutil.copy(os.path.join(all_path_root, f'{this_searched_name}反.png'), os.path.join(save_path, f'{this_searched_name}反.png'))
                                        shutil.copy(os.path.join(all_path_root, f'{this_searched_name}.info'), os.path.join(save_path, f'{this_searched_name}.info'))
                            #名字通过，开始检查地址
                            if name_passed_flag:
                                ws[f"{letter}{i}"] = check_r
                                with open(os.path.join(found_path, f'{check_r}.info'), 'r', encoding='utf-8') as reader:
                                    line = reader.readline()
                                info = json.loads(line.strip())
                                address = ws[f"{address_column_letters}{i}"].value
                                #香港人需要地址
                                if info['籍贯'] == '香港':
                                    if pd.isna(address) or address.strip().replace(' ', '') == '':
                                        ws[f"{address_column_letters}{i}"].font = red_font
                                        passed_flag = False
                                        error_rows = chech_and_add(error_rows, i)
                                        ws[f"{address_column_letters}{i}"] = '未填写地址'
                                #再检查如果有地址是否有错
                                if not pd.isna(address):
                                    if address.strip().replace(' ', '') != '':
                                        address = column_names['地址'](address)
                                        if pd.isna(address):
                                            ws[f"{address_column_letters}{i}"].font = red_font
                                            passed_flag = False
                                            error_rows = chech_and_add(error_rows, i)
                                        else:
                                            ws[f"{address_column_letters}{i}"] = address
                                    else:
                                        ws[f"{address_column_letters}{i}"] = np.nan
                    else:
                        ws[f"{letter}{i}"] = np.nan
    wb.save(file_path)
    return passed_flag, [str(i) for i in error_rows], extra_searched

def check_excel_after(file_path, sheet_name = None):
    passed_flag = True
    error_rows = []
    # 设置绿色字体
    red_font = Font(color="FF0000")  # 红色字体
    black_font = Font(color="000000")  # 黑色字体
    wb = load_workbook(file_path)
    sheets = []
    if sheet_name == None:
        sheets = sheet_names
    elif isinstance(sheet_name, str):
        if sheet_name in sheet_names:
            sheets = [sheet_name]
    elif isinstance(sheet_name, list):
        sheets = [i for i in sheet_name if i in sheet_names]
    for sheet in sheets:
        ws = wb[sheet]
        ws = read_openpyxl_by_str(ws)
        row_count = ws.max_row
        check_cloumns = [i for i in [cell.value for cell in ws[1]] if i in catch_columns]
        column_letters = get_column_letter(ws, check_cloumns + ['姓名'])
        name_letter = column_letters[-1]
        column_letters = column_letters[:-1]
        nation_index = check_cloumns.index('归属地')
        nation_letter = column_letters[nation_index]
        #检查一次拿一些行需要检查
        check_columns_log = []
        for i in range(2, row_count + 1):
            value = ws[f"{name_letter}{i}"].value
            if not pd.isna(value) and value.strip().replace(' ', ''):
                check_columns_log.append(i)
        for index, letter in enumerate(column_letters):
            if letter == nation_letter:
                nation_pass_flag = True
            else:
                nation_pass_flag = False
            c_name = check_cloumns[index]
            for i in check_columns_log:
                value = ws[f"{letter}{i}"].value
                ws[f"{letter}{i}"].font = black_font
                if pd.isna(value) or not value.strip().replace(' ', ''):
                    value = '内容缺失'
                    ws[f"{letter}{i}"].value = value
                    if nation_pass_flag:
                        ws[f"{letter}{i}"].font = red_font
                        passed_flag = False
                        error_rows = chech_and_add(error_rows, i)
                        continue
                elif nation_pass_flag:
                    if value not in ['香港', '内地']:
                        ws[f"{letter}{i}"].font = red_font
                        passed_flag = False
                        error_rows = chech_and_add(error_rows, i)
                    continue
                check_r = column_names_after[c_name](value)
                if check_r == None:
                    if c_name == '民族' and value == '无':
                        nation_value = ws[f"{nation_letter}{i}"].value
                        if nation_value != '香港':
                            ws[f"{letter}{i}"].font = red_font
                            passed_flag = False
                            error_rows = chech_and_add(error_rows, i)
                    else:
                        ws[f"{letter}{i}"].font = red_font
                        passed_flag = False
                        error_rows = chech_and_add(error_rows, i)
    wb.save(file_path)
    return passed_flag, [str(i) for i in error_rows]